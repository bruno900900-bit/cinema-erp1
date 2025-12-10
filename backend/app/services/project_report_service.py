"""
Serviço de Relatório de Projeto
Gera relatórios detalhados com informações do projeto e suas locações
"""
from typing import Dict, Any, List, Optional
from datetime import date, datetime
from sqlalchemy.orm import Session, joinedload
from io import BytesIO

from ..models.project import Project, ProjectStatus
from ..models.project_location import ProjectLocation, RentalStatus
from ..models.project_location_stage import ProjectLocationStage, LocationStageType, StageStatus


class ProjectReportService:
    def __init__(self, db: Session):
        self.db = db

    def _format_date(self, d: Optional[date]) -> str:
        """Formata uma data para exibição ou retorna '-' se nula"""
        if d is None:
            return "-"
        return d.strftime("%d/%m/%Y")

    def _format_datetime(self, dt: Optional[datetime]) -> str:
        """Formata datetime para exibição"""
        if dt is None:
            return "-"
        return dt.strftime("%d/%m/%Y %H:%M")

    def _get_status_label(self, status: RentalStatus) -> str:
        """Retorna label legível para status de locação"""
        labels = {
            RentalStatus.RESERVED: "Reservada",
            RentalStatus.CONFIRMED: "Confirmada",
            RentalStatus.IN_USE: "Em Uso",
            RentalStatus.RETURNED: "Devolvida",
            RentalStatus.OVERDUE: "Atrasada",
            RentalStatus.CANCELLED: "Cancelada",
        }
        return labels.get(status, str(status.value))

    def _get_project_status_label(self, status: ProjectStatus) -> str:
        """Retorna label legível para status do projeto"""
        labels = {
            ProjectStatus.ACTIVE: "Ativo",
            ProjectStatus.ARCHIVED: "Arquivado",
            ProjectStatus.COMPLETED: "Concluído",
            ProjectStatus.ON_HOLD: "Em Espera",
            ProjectStatus.CANCELLED: "Cancelado",
        }
        return labels.get(status, str(status.value))

    def get_project_report(self, project_id: int) -> Dict[str, Any]:
        """Gera relatório completo do projeto"""
        project = self.db.query(Project).options(
            joinedload(Project.project_locations).joinedload(ProjectLocation.location),
            joinedload(Project.project_locations).joinedload(ProjectLocation.stages),
        ).filter(Project.id == project_id).first()

        if not project:
            return None

        # Informações do projeto
        project_info = {
            "id": project.id,
            "nome": project.title or project.name,
            "descricao": project.description or "-",
            "cliente": project.client_name or "-",
            "cliente_email": project.client_email or "-",
            "cliente_telefone": project.client_phone or "-",
            "status": self._get_project_status_label(project.status),
            "data_inicio": self._format_date(project.start_date),
            "data_fim": self._format_date(project.end_date),
            "orcamento_total": project.budget_total or 0,
            "orcamento_gasto": project.budget_spent or 0,
            "orcamento_restante": project.budget_remaining or 0,
            "moeda": project.budget_currency or "BRL",
            "criado_em": self._format_datetime(project.created_at) if project.created_at else "-",
        }

        # Calcular totais
        total_locacoes = len(project.project_locations)
        total_custo_locacoes = sum(pl.total_cost or 0 for pl in project.project_locations)
        locacoes_ativas = sum(1 for pl in project.project_locations if pl.status in [RentalStatus.CONFIRMED, RentalStatus.IN_USE])
        locacoes_concluidas = sum(1 for pl in project.project_locations if pl.status == RentalStatus.RETURNED)

        # Informações das locações
        locacoes = []
        for pl in project.project_locations:
            location = pl.location

            # Período de locação formatado
            periodo_locacao = f"{self._format_date(pl.rental_start)} até {self._format_date(pl.rental_end)}"

            # Período de filmagem
            if pl.filming_start_date and pl.filming_end_date:
                periodo_filmagem = f"{self._format_date(pl.filming_start_date)} até {self._format_date(pl.filming_end_date)}"
            elif pl.filming_start_date:
                periodo_filmagem = self._format_date(pl.filming_start_date)
            else:
                periodo_filmagem = "-"

            locacao_info = {
                "id": pl.id,
                "nome": location.title if location else f"Locação #{pl.location_id}",
                "cidade": location.city if location else "-",
                "estado": location.state if location else "-",
                "endereco": f"{location.street or ''}, {location.number or ''}".strip(", ") if location else "-",
                "valor_diaria": pl.daily_rate or 0,
                "valor_hora": pl.hourly_rate,
                "valor_total": pl.total_cost or 0,
                "moeda": pl.currency or "BRL",
                "periodo_locacao": periodo_locacao,
                "data_inicio": self._format_date(pl.rental_start),
                "data_fim": self._format_date(pl.rental_end),
                "duracao_dias": pl.duration_days,
                "data_visita": self._format_date(pl.visit_date),
                "data_visita_tecnica": self._format_date(pl.technical_visit_date),
                "periodo_filmagem": periodo_filmagem,
                "data_entrega": self._format_date(pl.delivery_date),
                "status": self._get_status_label(pl.status),
                "progresso": round(pl.completion_percentage or 0, 1),
                "observacoes": pl.notes or "-",
                "requisitos_especiais": pl.special_requirements or "-",
                "equipamentos": pl.equipment_needed or "-",
            }
            locacoes.append(locacao_info)

        # Ordenar por data de início
        locacoes.sort(key=lambda x: x.get("data_inicio", ""))

        return {
            "projeto": project_info,
            "resumo": {
                "total_locacoes": total_locacoes,
                "locacoes_ativas": locacoes_ativas,
                "locacoes_concluidas": locacoes_concluidas,
                "custo_total_locacoes": total_custo_locacoes,
            },
            "locacoes": locacoes,
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        }

    def export_to_excel(self, project_id: int) -> Optional[BytesIO]:
        """Exporta relatório para Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl é necessário para exportar Excel. Instale com: pip install openpyxl")

        report = self.get_project_report(project_id)
        if not report:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório do Projeto"

        # Estilos
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # ===== CABEÇALHO DO PROJETO =====
        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}'] = f"RELATÓRIO DO PROJETO: {report['projeto']['nome']}"
        ws[f'A{row}'].font = Font(bold=True, size=16, color="4F46E5")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Informações do Projeto
        ws[f'A{row}'] = "INFORMAÇÕES DO PROJETO"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        info_fields = [
            ("Cliente:", report['projeto']['cliente']),
            ("Status:", report['projeto']['status']),
            ("Período:", f"{report['projeto']['data_inicio']} até {report['projeto']['data_fim']}"),
            ("Orçamento Total:", f"R$ {report['projeto']['orcamento_total']:,.2f}"),
            ("Orçamento Gasto:", f"R$ {report['projeto']['orcamento_gasto']:,.2f}"),
            ("Orçamento Restante:", f"R$ {report['projeto']['orcamento_restante']:,.2f}"),
        ]

        for label, value in info_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # ===== RESUMO =====
        ws[f'A{row}'] = "RESUMO DAS LOCAÇÕES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        resumo_fields = [
            ("Total de Locações:", report['resumo']['total_locacoes']),
            ("Locações Ativas:", report['resumo']['locacoes_ativas']),
            ("Locações Concluídas:", report['resumo']['locacoes_concluidas']),
            ("Custo Total:", f"R$ {report['resumo']['custo_total_locacoes']:,.2f}"),
        ]

        for label, value in resumo_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 2

        # ===== TABELA DE LOCAÇÕES =====
        ws[f'A{row}'] = "DETALHAMENTO DAS LOCAÇÕES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:L{row}')
        row += 1

        # Cabeçalhos da tabela
        headers = [
            "Nome da Locação",
            "Cidade",
            "Valor Diária (R$)",
            "Valor Total (R$)",
            "Período de Locação",
            "Data da Visita",
            "Visita Técnica",
            "Período Filmagem",
            "Data Entrega",
            "Status",
            "Progresso (%)",
            "Observações"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row += 1

        # Dados das locações
        for loc in report['locacoes']:
            data = [
                loc['nome'],
                loc['cidade'],
                loc['valor_diaria'],
                loc['valor_total'],
                loc['periodo_locacao'],
                loc['data_visita'],
                loc['data_visita_tecnica'],
                loc['periodo_filmagem'],
                loc['data_entrega'],
                loc['status'],
                loc['progresso'],
                loc['observacoes'],
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Ajustar largura das colunas
        column_widths = [30, 15, 15, 15, 25, 15, 15, 25, 15, 15, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Rodapé
        row += 2
        ws[f'A{row}'] = f"Relatório gerado em: {report['gerado_em']}"
        ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")

        # Salvar em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
